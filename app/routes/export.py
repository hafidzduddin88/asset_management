# app/routes/export.py
from fastapi import APIRouter, Request, Depends, Form
from fastapi.responses import StreamingResponse
from fastapi.templating import Jinja2Templates
from app.utils.auth import get_current_profile
from app.utils.database_manager import get_supabase
from app.utils.device_detector import get_template
import io
import json
from datetime import datetime
from typing import List
from collections import OrderedDict

router = APIRouter(prefix="/export", tags=["export"])
templates = Jinja2Templates(directory="app/templates")

# Available tables for export
EXPORT_TABLES = {
    'assets': {
        'name': 'Assets',
        'table': 'assets',
        'columns': OrderedDict([
            ('asset_tag', 'Asset Tag'),
            ('asset_name', 'Asset Name'),
            ('status', 'Status'),
            ('item_condition', 'Condition'),
            ('category_name', 'Category'),
            ('type_name', 'Asset Type'),
            ('manufacture', 'Manufacture'),
            ('model', 'Model'),
            ('serial_number', 'Serial Number'),
            ('company_name', 'Company'),
            ('business_unit_name', 'Business Unit'),
            ('owner_name', 'Owner'),
            ('location_name', 'Location'),
            ('room_name', 'Room'),
            ('purchase_date', 'Purchase Date'),
            ('purchase_cost', 'Purchase Cost (Rp)'),
            ('book_value', 'Book Value (Rp)'),
            ('depreciation_value', 'Depreciation Value (Rp)'),
            ('residual_value', 'Residual Value (Rp)'),
            ('supplier', 'Supplier'),
            ('warranty', 'Warranty'),
            ('notes', 'Notes'),
            ('asset_id', 'Asset ID')
        ])
    },
    'approvals': {
        'name': 'Approvals',
        'table': 'approvals',
        'columns': OrderedDict([
            ('type', 'Type'),
            ('asset_name', 'Asset Name'),
            ('status', 'Status'),
            ('submitted_by', 'Submitted By'),
            ('submitted_date', 'Submitted Date'),
            ('approved_by', 'Approved By'),
            ('approved_date', 'Approved Date'),
            ('description', 'Description'),
            ('approval_id', 'Approval ID')
        ])
    },
    'damage_log': {
        'name': 'Damage Log',
        'table': 'damage_log',
        'columns': OrderedDict([
            ('asset_name', 'Asset Name'),
            ('damage_type', 'Damage Type'),
            ('severity', 'Severity'),
            ('status', 'Status'),
            ('description', 'Description'),
            ('reported_by', 'Reported By'),
            ('report_date', 'Report Date'),
            ('damage_id', 'Damage ID')
        ])
    },
    'repair_log': {
        'name': 'Repair Log',
        'table': 'repair_log',
        'columns': OrderedDict([
            ('asset_name', 'Asset Name'),
            ('repair_action', 'Repair Action'),
            ('status', 'Status'),
            ('description', 'Description'),
            ('performed_by', 'Performed By'),
            ('repair_date', 'Repair Date'),
            ('repair_id', 'Repair ID')
        ])
    },
    'users': {
        'name': 'Users',
        'table': 'profiles',
        'columns': OrderedDict([
            ('full_name', 'Full Name'),
            ('username', 'Email'),
            ('role', 'Role'),
            ('business_unit_name', 'Business Unit'),
            ('is_active', 'Active Status'),
            ('email_verified', 'Email Verified'),
            ('created_at', 'Created Date'),
            ('last_login_at', 'Last Login'),
            ('id', 'User ID')
        ])
    }
}

@router.get("/")
async def export_page(request: Request, current_profile=Depends(get_current_profile)):
    """Export data page - accessible by all users"""
    # Filter tables based on user role
    available_tables = EXPORT_TABLES.copy()
    if current_profile.role not in ['admin']:
        # Non-admin users cannot export user data
        available_tables.pop('users', None)
    
    template_path = get_template(request, "export/index.html")
    return templates.TemplateResponse(template_path, {
        "request": request,
        "user": current_profile,
        "tables": available_tables
    })

@router.post("/excel")
async def export_to_excel(
    request: Request,
    table: str = Form(...),
    columns: List[str] = Form(...),
    exclude_disposed: bool = Form(False),
    exclude_to_be_disposed: bool = Form(False),
    exclude_damaged: bool = Form(False),
    current_profile=Depends(get_current_profile)
):
    """Export selected table and columns to Excel - accessible by all users"""
    try:
        # Import openpyxl here to avoid dependency if not used
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        if table not in EXPORT_TABLES:
            raise ValueError("Invalid table selected")
        
        # Check user permissions for table access
        if table == 'users' and current_profile.role not in ['admin']:
            raise ValueError("Access denied: Admin role required for user data export")
        
        table_config = EXPORT_TABLES[table]
        supabase = get_supabase()
        
        # Build query with foreign key relationships for assets
        if table == 'assets':
            # Include foreign key relationships for proper data display
            select_fields = []
            for col in columns:
                if col in ['asset_id', 'asset_name', 'asset_tag', 'manufacture', 'model', 'serial_number', 'purchase_date', 'purchase_cost', 'status', 'item_condition', 'room_name', 'notes', 'warranty', 'supplier', 'depreciation_value', 'residual_value', 'book_value']:
                    select_fields.append(col)
            
            # Add foreign key relationships
            select_fields.extend([
                'ref_categories(category_name)',
                'ref_asset_types(type_name)', 
                'ref_locations(location_name, room_name)',
                'ref_business_units(business_unit_name)',
                'ref_companies(company_name)',
                'ref_owners(owner_name)'
            ])
            
            query = supabase.table(table_config['table']).select(','.join(select_fields))
        elif table == 'users':
            # Include business unit relationship for users
            select_fields = []
            for col in columns:
                if col in ['id', 'username', 'full_name', 'role', 'is_active', 'email_verified', 'created_at', 'last_login_at', 'business_unit_name']:
                    select_fields.append(col)
            
            select_fields.append('ref_business_units(business_unit_name)')
            query = supabase.table(table_config['table']).select(','.join(select_fields))
        else:
            query = supabase.table(table_config['table']).select(','.join(columns))
        
        # Apply filters
        if table == 'assets':
            if exclude_disposed:
                query = query.neq('status', 'Disposed')
            if exclude_to_be_disposed:
                query = query.neq('status', 'To be Disposed')
            if exclude_damaged:
                query = query.neq('status', 'Damaged')
        
        # Apply sorting
        if table == 'assets':
            query = query.order('asset_tag')
        elif table == 'approvals':
            query = query.order('submitted_date', desc=True)
        elif table == 'damage_log':
            query = query.order('created_at', desc=True)
        elif table == 'repair_log':
            query = query.order('created_at', desc=True)
        elif table == 'users':
            query = query.order('full_name')
        
        # Execute query
        response = query.execute()
        data = response.data
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = table_config['name']
        
        # Add headers in the order they appear in the form
        ordered_columns = [col for col in table_config['columns'].keys() if col in columns]
        headers = [table_config['columns'][col] for col in ordered_columns]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data rows in correct order
        for row in data:
            row_data = []
            for col in ordered_columns:
                value = row.get(col, '')
                
                # Handle foreign key relationships
                if table == 'assets':
                    if col == 'category_name':
                        value = row.get('ref_categories', {}).get('category_name', '') if row.get('ref_categories') else ''
                    elif col == 'type_name':
                        value = row.get('ref_asset_types', {}).get('type_name', '') if row.get('ref_asset_types') else ''
                    elif col == 'location_name':
                        value = row.get('ref_locations', {}).get('location_name', '') if row.get('ref_locations') else ''
                    elif col == 'business_unit_name':
                        value = row.get('ref_business_units', {}).get('business_unit_name', '') if row.get('ref_business_units') else ''
                    elif col == 'company_name':
                        value = row.get('ref_companies', {}).get('company_name', '') if row.get('ref_companies') else ''
                    elif col == 'owner_name':
                        value = row.get('ref_owners', {}).get('owner_name', '') if row.get('ref_owners') else ''
                elif table == 'users':
                    if col == 'business_unit_name':
                        value = row.get('ref_business_units', {}).get('business_unit_name', '') if row.get('ref_business_units') else ''
                    elif col == 'is_active':
                        value = 'Active' if value else 'Inactive'
                    elif col == 'email_verified':
                        value = 'Verified' if value else 'Not Verified'
                
                # Format dates
                if col.endswith('_date') and value:
                    try:
                        if 'T' in str(value):
                            dt = datetime.fromisoformat(str(value).replace('Z', '+00:00'))
                            value = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass
                        
                row_data.append(value or '')
            ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{table_config['name']}_{timestamp}.xlsx"
        
        # Return as streaming response
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "Excel export not available - openpyxl not installed"}, status_code=500)
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": f"Export failed: {str(e)}"}, status_code=500)