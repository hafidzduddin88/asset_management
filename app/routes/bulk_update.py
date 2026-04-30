from fastapi import APIRouter, Request, Form, UploadFile, File, Depends
from fastapi.responses import RedirectResponse, StreamingResponse
from starlette.templating import Jinja2Templates
from app.utils.device_detector import get_template
from app.utils.auth import get_current_profile
from app.utils.database_manager import get_supabase, TABLES, invalidate_cache
from app.utils.flash import set_flash
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import logging
from datetime import datetime

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

@router.get("/bulk-update")
async def bulk_update_page(request: Request, current_profile = Depends(get_current_profile)):
    """Step 1: Export data with filters"""
    if current_profile.role != "admin":
        return RedirectResponse("/", status_code=303)
    
    try:
        supabase = get_supabase()
        
        # Get filter options
        categories = supabase.table('ref_categories').select('category_name').execute().data
        asset_types = supabase.table('ref_asset_types').select('type_name').execute().data
        locations = supabase.table('ref_locations').select('location_name, room_name').execute().data
        owners = supabase.table('ref_owners').select('owner_name').execute().data
        
        # Group locations by location_name
        location_dict = {}
        for loc in locations:
            location_name = loc.get('location_name')
            room_name = loc.get('room_name')
            if location_name:
                if location_name not in location_dict:
                    location_dict[location_name] = []
                if room_name:
                    location_dict[location_name].append(room_name)
        
        template_path = get_template(request, "bulk_update/index.html")
        return templates.TemplateResponse(template_path, {
            "request": request,
            "user": current_profile,
            "categories": [c['category_name'] for c in categories],
            "asset_types": [t['type_name'] for t in asset_types],
            "locations": location_dict,
            "owners": [o['owner_name'] for o in owners],
            "statuses": ["Active", "Damaged", "Lost", "Disposed"]
        })
    except Exception as e:
        logging.error(f"Error loading bulk update page: {str(e)}")
        return RedirectResponse("/", status_code=303)

@router.post("/bulk-update/export")
async def bulk_update_export(
    request: Request,
    current_profile = Depends(get_current_profile),
    category: str = Form(None),
    asset_type: str = Form(None),
    location: str = Form(None),
    room: str = Form(None),
    owner: str = Form(None),
    owner_type: str = Form(None),
    status: str = Form(None)
):
    """Step 1: Export filtered assets to Excel"""
    if current_profile.role != "admin":
        return RedirectResponse("/", status_code=303)
    
    try:
        supabase = get_supabase()
        
        # Build query with filters
        query = supabase.table(TABLES['ASSETS']).select('''
            asset_id, asset_name, manufacture, model, serial_number, asset_tag,
            room_name, notes, item_condition, purchase_date, purchase_cost,
            warranty, supplier, journal, status, year,
            ref_categories(category_name),
            ref_asset_types(type_name),
            ref_locations(location_name, room_name),
            ref_owners(owner_name),
            ref_companies(company_name),
            ref_business_units(business_unit_name)
        ''')
        
        # Apply filters
        if category:
            cat_response = supabase.table('ref_categories').select('category_id').eq('category_name', category).execute()
            if cat_response.data:
                query = query.eq('category_id', cat_response.data[0]['category_id'])
        
        if asset_type:
            type_response = supabase.table('ref_asset_types').select('asset_type_id').eq('type_name', asset_type).execute()
            if type_response.data:
                query = query.eq('asset_type_id', type_response.data[0]['asset_type_id'])
        
        if location and room:
            loc_response = supabase.table('ref_locations').select('location_id').eq('location_name', location).eq('room_name', room).execute()
            if loc_response.data:
                query = query.eq('location_id', loc_response.data[0]['location_id'])
        
        if owner:
            owner_response = supabase.table('ref_owners').select('owner_id').eq('owner_name', owner).execute()
            if owner_response.data:
                query = query.eq('owner_id', owner_response.data[0]['owner_id'])
        
        if owner_type:
            query = query.eq('owner_type', owner_type)
        
        if status:
            query = query.eq('status', status)
        
        response = query.execute()
        assets = response.data
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Assets"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = [
            "asset_id", "asset_name", "category_name", "type_name", "manufacture", 
            "model", "serial_number", "asset_tag", "company_name", "business_unit_name",
            "location_name", "room_name", "owner_name", "owner_type", "assigned_user_name",
            "item_condition", "purchase_date", "purchase_cost", "warranty", "supplier", 
            "journal", "notes", "status", "year"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_num, asset in enumerate(assets, 2):
            ws.cell(row=row_num, column=1, value=asset.get('asset_id'))
            ws.cell(row=row_num, column=2, value=asset.get('asset_name'))
            ws.cell(row=row_num, column=3, value=asset.get('ref_categories', {}).get('category_name') if asset.get('ref_categories') else '')
            ws.cell(row=row_num, column=4, value=asset.get('ref_asset_types', {}).get('type_name') if asset.get('ref_asset_types') else '')
            ws.cell(row=row_num, column=5, value=asset.get('manufacture'))
            ws.cell(row=row_num, column=6, value=asset.get('model'))
            ws.cell(row=row_num, column=7, value=asset.get('serial_number'))
            ws.cell(row=row_num, column=8, value=asset.get('asset_tag'))
            ws.cell(row=row_num, column=9, value=asset.get('ref_companies', {}).get('company_name') if asset.get('ref_companies') else '')
            ws.cell(row=row_num, column=10, value=asset.get('ref_business_units', {}).get('business_unit_name') if asset.get('ref_business_units') else '')
            
            loc_data = asset.get('ref_locations', {})
            ws.cell(row=row_num, column=11, value=loc_data.get('location_name') if loc_data else '')
            ws.cell(row=row_num, column=12, value=asset.get('room_name'))
            ws.cell(row=row_num, column=13, value=asset.get('ref_owners', {}).get('owner_name') if asset.get('ref_owners') else '')
            ws.cell(row=row_num, column=14, value=asset.get('owner_type', 'GA'))
            
            # Assigned user name for IT assets
            if asset.get('owner_type') == 'IT':
                # Use stored assigned_user_name if available, otherwise get from user object
                assigned_user_name = asset.get('assigned_user_name')
                if not assigned_user_name:
                    assigned_user = asset.get('assigned_user', {})
                    assigned_user_name = assigned_user.get('full_name') if assigned_user else ''
            else:
                assigned_user_name = ''
            
            ws.cell(row=row_num, column=15, value=assigned_user_name)
            
            ws.cell(row=row_num, column=16, value=asset.get('item_condition'))
            ws.cell(row=row_num, column=17, value=asset.get('purchase_date'))
            ws.cell(row=row_num, column=18, value=asset.get('purchase_cost'))
            ws.cell(row=row_num, column=19, value=asset.get('warranty'))
            ws.cell(row=row_num, column=20, value=asset.get('supplier'))
            ws.cell(row=row_num, column=21, value=asset.get('journal'))
            ws.cell(row=row_num, column=22, value=asset.get('notes'))
            ws.cell(row=row_num, column=23, value=asset.get('status'))
            ws.cell(row=row_num, column=24, value=asset.get('year'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"assets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logging.error(f"Error exporting assets: {str(e)}")
        response = RedirectResponse("/bulk-update", status_code=303)
        set_flash(response, f"Gagal export data: {str(e)}", "error")
        return response

@router.post("/bulk-update/import")
async def bulk_update_import(
    request: Request,
    current_profile = Depends(get_current_profile),
    file: UploadFile = File(...)
):
    """Step 2: Import Excel and show preview"""
    if current_profile.role != "admin":
        return RedirectResponse("/", status_code=303)
    
    try:
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Parse data
        headers = [cell.value for cell in ws[1]]
        updates = []
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:  # Skip if no asset_id
                continue
            
            try:
                update_data = {
                    'asset_id': row[0],
                    'asset_name': row[1],
                    'category_name': row[2],
                    'type_name': row[3],
                    'manufacture': row[4],
                    'model': row[5],
                    'serial_number': row[6],
                    'asset_tag': row[7],
                    'company_name': row[8],
                    'business_unit_name': row[9],
                    'location_name': row[10],
                    'room_name': row[11],
                    'owner_name': row[12],
                    'owner_type': row[13] if len(row) > 13 else 'GA',
                    'assigned_user_name': row[14] if len(row) > 14 else None,
                    'item_condition': row[15] if len(row) > 15 else row[13],
                    'purchase_date': str(row[16]) if len(row) > 16 and row[16] else str(row[14]) if row[14] else None,
                    'purchase_cost': float(row[17]) if len(row) > 17 and row[17] else float(row[15]) if row[15] else None,
                    'warranty': row[18] if len(row) > 18 else row[16],
                    'supplier': row[19] if len(row) > 19 else row[17],
                    'journal': row[20] if len(row) > 20 else row[18],
                    'notes': row[21] if len(row) > 21 else row[19],
                    'status': row[22] if len(row) > 22 else row[20],
                    'year': int(row[23]) if len(row) > 23 and row[23] else int(row[21]) if row[21] else None
                }
                updates.append(update_data)
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        # Store in session (simplified - use database or cache in production)
        request.session['bulk_updates'] = updates
        
        template_path = get_template(request, "bulk_update/preview.html")
        return templates.TemplateResponse(template_path, {
            "request": request,
            "user": current_profile,
            "updates": updates,
            "errors": errors,
            "total": len(updates)
        })
        
    except Exception as e:
        logging.error(f"Error importing file: {str(e)}")
        response = RedirectResponse("/bulk-update", status_code=303)
        set_flash(response, f"Gagal import file: {str(e)}", "error")
        return response

@router.post("/bulk-update/confirm")
async def bulk_update_confirm(request: Request, current_profile = Depends(get_current_profile)):
    """Step 3: Execute bulk update"""
    if current_profile.role != "admin":
        return RedirectResponse("/", status_code=303)
    
    try:
        updates = request.session.get('bulk_updates', [])
        if not updates:
            response = RedirectResponse("/bulk-update", status_code=303)
            set_flash(response, "Tidak ada data untuk diupdate", "error")
            return response
        
        supabase = get_supabase()
        success_count = 0
        error_count = 0
        errors = []
        
        for update_data in updates:
            try:
                asset_id = update_data['asset_id']
                
                # Prepare update payload
                payload = {}
                
                # Resolve foreign keys
                if update_data.get('category_name'):
                    cat = supabase.table('ref_categories').select('category_id').eq('category_name', update_data['category_name']).execute()
                    if cat.data:
                        payload['category_id'] = cat.data[0]['category_id']
                
                if update_data.get('type_name'):
                    typ = supabase.table('ref_asset_types').select('asset_type_id').eq('type_name', update_data['type_name']).execute()
                    if typ.data:
                        payload['asset_type_id'] = typ.data[0]['asset_type_id']
                
                if update_data.get('company_name'):
                    comp = supabase.table('ref_companies').select('company_id').eq('company_name', update_data['company_name']).execute()
                    if comp.data:
                        payload['company_id'] = comp.data[0]['company_id']
                
                if update_data.get('business_unit_name'):
                    bu = supabase.table('ref_business_units').select('business_unit_id').eq('business_unit_name', update_data['business_unit_name']).execute()
                    if bu.data:
                        payload['business_unit_id'] = bu.data[0]['business_unit_id']
                
                if update_data.get('location_name') and update_data.get('room_name'):
                    loc = supabase.table('ref_locations').select('location_id').eq('location_name', update_data['location_name']).eq('room_name', update_data['room_name']).execute()
                    if loc.data:
                        payload['location_id'] = loc.data[0]['location_id']
                
                if update_data.get('owner_name'):
                    own = supabase.table('ref_owners').select('owner_id').eq('owner_name', update_data['owner_name']).execute()
                    if own.data:
                        payload['owner_id'] = own.data[0]['owner_id']
                
                # Handle assigned user for IT assets
                if update_data.get('owner_type') == 'IT' and update_data.get('assigned_user_name'):
                    user_name = update_data['assigned_user_name']
                    # Try to find user by full_name first, then username
                    user = supabase.table('profiles').select('id').eq('full_name', user_name).execute()
                    if not user.data:
                        user = supabase.table('profiles').select('id').eq('username', user_name).execute()
                    
                    if user.data:
                        payload['assigned_user_id'] = user.data[0]['id']
                        payload['assigned_user_name'] = user_name
                    else:
                        errors.append(f"Asset ID {update_data.get('asset_id')}: User '{user_name}' not found")
                        continue
                
                # Direct fields
                direct_fields = ['asset_name', 'manufacture', 'model', 'serial_number', 'asset_tag',
                                'room_name', 'item_condition', 'purchase_date', 'purchase_cost',
                                'warranty', 'supplier', 'journal', 'notes', 'status', 'year', 'owner_type', 'assigned_user_name']
                
                for field in direct_fields:
                    if update_data.get(field) is not None:
                        payload[field] = update_data[field]
                
                # Update asset
                supabase.table(TABLES['ASSETS']).update(payload).eq('asset_id', asset_id).execute()
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"Asset ID {update_data.get('asset_id')}: {str(e)}")
                logging.error(f"Error updating asset {update_data.get('asset_id')}: {str(e)}")
        
        # Clear session
        request.session.pop('bulk_updates', None)
        
        # Invalidate cache
        invalidate_cache()
        
        # Show result
        template_path = get_template(request, "bulk_update/result.html")
        return templates.TemplateResponse(template_path, {
            "request": request,
            "user": current_profile,
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors,
            "total": len(updates)
        })
        
    except Exception as e:
        logging.error(f"Error confirming bulk update: {str(e)}")
        response = RedirectResponse("/bulk-update", status_code=303)
        set_flash(response, f"Gagal update data: {str(e)}", "error")
        return response
